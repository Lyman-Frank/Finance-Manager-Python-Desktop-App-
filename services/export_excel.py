from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import PieChart, Reference

from database.crud import get_transactions_filtered
from config.constants import TRANSACTION_TYPES_REVERSE

def export_transactions_to_excel(file_path: str,
                                 transaction_type=None,
                                 date_from=None,
                                 date_to=None):

    data = get_transactions_filtered(
        transaction_type=transaction_type,
        date_from=date_from,
        date_to=date_to
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Операции"

    headers = ["Дата", "Тип", "Категория", "Сумма", "Комментарий"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    total_income = 0.0
    total_expense = 0.0

    income_fill = PatternFill(
        start_color="a5d6a8",
        end_color="a5d6a8",
        fill_type="solid"
    )

    expense_fill = PatternFill(
        start_color="f2978d",
        end_color="f2978d",
        fill_type="solid"
    )

    for row in data:
        amount = float(row["amount"])

        if row["type"] == "income":
            total_income += amount
            fill = income_fill
        else:
            total_expense += amount
            fill = expense_fill

        ws.append([
            row["date"],
            TRANSACTION_TYPES_REVERSE.get(row["type"], row["type"]),
            row["category"],
            amount,
            row["comment"]
        ])

        current_row = ws.max_row

        for col in range(1, 6):
            ws.cell(row=current_row, column=col).fill = fill

    # ----- TOTALS -----
    ws.append([])
    ws.append(["ИТОГО"])
    ws.append(["Доходы", round(total_income, 2)])
    ws.append(["Расходы", round(total_expense, 2)])

    ws["A" + str(ws.max_row - 2)].font = Font(bold=True)
    ws["A" + str(ws.max_row - 1)].font = Font(bold=True)
    ws["A" + str(ws.max_row)].font = Font(bold=True)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # ================= SUMMARY SHEET =================
    summary = wb.create_sheet("Сводка")

    summary["A1"] = "Сводка по операциям"
    summary["A1"].font = Font(bold=True)

    summary["A3"] = "Доходы"
    summary["B3"] = round(total_income, 2)

    summary["A4"] = "Расходы"
    summary["B4"] = round(total_expense, 2)

    summary["A5"] = "Баланс"
    summary["B5"] = round(total_income - total_expense, 2)

    summary["A3"].font = Font(bold=True)
    summary["A4"].font = Font(bold=True)
    summary["A5"].font = Font(bold=True)

    chart = PieChart()
    chart.title = "Доходы и расходы"

    labels = Reference(summary, min_col=1, min_row=3, max_row=4)
    data = Reference(summary, min_col=2, min_row=3, max_row=4)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    chart.height = 10
    chart.width = 10

    summary.add_chart(chart, "D3")

    wb.save(file_path)
